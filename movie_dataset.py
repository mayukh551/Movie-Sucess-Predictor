# Using web scraping to find
# last 5 movies performed by an actor or director

from openpyxl import workbook, load_workbook

wb = load_workbook('Trial.xlsx')
ws = wb.active
r = 37
row = 'A' + str(r)
popular_actors = gd_actors = avg_actors = flop_actors = 0
seven_pointer = {
    'popular_actors': 0, 'gd_actors': 0, 'avg_actors': 0, 'flop_actors': 0}

six_pointer = {
    'popular_actors': 0, 'gd_actors': 0, 'avg_actors': 0, 'flop_actors': 0}

# director_performance = {
#     'popular_directors': 0, 'ok_directors': 0, 'flop_directors': 0
# }

count = 0
dp = dok = dflop = 0
while ws[row].value is not None:
    if ws['H' + str(r)].value == 1:
        dp += 1

    elif ws['I' + str(r)].value == 1:
        dok += 1

    elif ws['J' + str(r)].value == 1:
        dflop += 1

    if ws['H' + str(r)].value == 0 and ws['I' + str(r)].value == 0 and ws['J' + str(r)].value == 0:
        print("Director Not found; Row no. : ", r)
        count = count - 1

    count += 1
    r = r + 1
    row = 'A' + str(r)

print('For average Movies')
print(f'Popular Director : {dp} out of {count}       -> {(dp / count) * 100}%')
print(f'Average Director : {dok} out of {count}      -> {(dok / count) * 100}%')
print(f'Flop Director    : {dflop} out of {count}       -> {(dflop / count) * 100}%')
print('\n')
r = 37
row = 'A' + str(r)
count = 0
while ws[row].value is not None:
    if 7 <= float(ws['C' + str(r)].value) <= 7.7:
        # print(">=7")
        # if (int(ws['D' + str(r)].value) + int(ws['E' + str(r)].value) + int(ws['F' + str(r)].value) + int(
        #         ws['G' + str(r)].value)):
        seven_pointer['popular_actors'] += int(ws['D' + str(r)].value)
        seven_pointer['gd_actors'] += int(ws['E' + str(r)].value)
        seven_pointer['avg_actors'] += int(ws['F' + str(r)].value)
        seven_pointer['flop_actors'] += int(ws['G' + str(r)].value)

    # elif 6 <= float(ws['C' + str(r)].value) < 7:
    #     # print(">=6")
    #     six_pointer['popular_actors'] += int(ws['D' + str(r)].value)
    #     six_pointer['gd_actors'] += int(ws['E' + str(r)].value)
    #     six_pointer['avg_actors'] += int(ws['F' + str(r)].value)
    #     six_pointer['flop_actors'] += int(ws['G' + str(r)].value)

    r = r + 1
    row = 'A' + str(r)
    count += 1

print('\nFor Movies : 7.7>= score >=7 out of', count, '\n')

print('Total Popular actors   :   ', seven_pointer["popular_actors"],
      f' ->  {((seven_pointer["popular_actors"] / count) * 100)}%')
print('Total good actors      :   ', seven_pointer["gd_actors"],
      f' ->  {((seven_pointer["gd_actors"] / count) * 100)}%')
print('Total average actors   :   ', seven_pointer["avg_actors"],
      f' ->  {((seven_pointer["avg_actors"] / count) * 100)}%')
print('Total flop actors      :   ', seven_pointer["flop_actors"],
      f' ->  {((seven_pointer["flop_actors"] / count) * 100)}%')

# print('\nFor Movies : 7> score >=6 out of', count, '\n')
#
# print('Total Popular actors   :   ', six_pointer["popular_actors"],
#       f'  ->   {((six_pointer["popular_actors"] / count) * 100)}%')
# print('Total good actors      :   ', six_pointer["gd_actors"],
#       f' ->  {((six_pointer["gd_actors"] / count) * 100)}%')
# print('Total average actors   :   ', six_pointer["avg_actors"],
#       f' ->  {((six_pointer["avg_actors"] / count) * 100)}%')
# print('Total flop actors      :   ', six_pointer["flop_actors"],
#       f' ->  {((six_pointer["flop_actors"] / count) * 100)}%')

wb.save('Trial.xlsx')
