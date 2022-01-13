# Using web scraping to find
# last 5 movies performed by an actor or director

from openpyxl import workbook, load_workbook
# from Get_Actor_List import *
# from Get_Cast_Score import *
# from Director_Movies_Score import *
from Movie_funcs import *
import requests

rel_year_of_movie = 0


def extract_best_movies(ratings):
    s = 0
    average = 0.0
    no_of_scores = len(ratings)
    # s is score out of 10
    # So is average
    if no_of_scores >= 4:
        s = ratings[-1] + ratings[-2] + ratings[-3]
        average = s / 3
    elif no_of_scores == 3:
        s = sum(ratings[1:])
        average = s / 2
    elif no_of_scores == 2:
        average = sum(ratings) / 2
    elif no_of_scores == 1:
        average = sum(ratings)
    return average


wb = load_workbook('Trial.xlsx')
ws = wb.active

# r = 37
# row = 'A' + str(r)
# movie_name = row
# rel_year = 'B' + str(r)
#
# # p = ok = flop = 0
# # dp = dok = dflop = 0
# popular_actors = gd_actors = avg_actors = flop_actors = 0
#
# count = 1
# while ws[movie_name].value is not None:
#     print('\n', ws[movie_name].value, '\n')
#     rel_year = 'B' + str(r)
#     cast_list = findCast(ws[movie_name].value, ws[rel_year].value)
#     score = []
#     actor_score = {}
#     # if cast List not empty
#     score_pref = {}
#     actor_performance = {}
#     print('Cast : ', cast_list)
#     for actor in cast_list:
#         # returns score list of past movies
#         score = cast_score_main_code(actor, ws['B' + str(r)].value)
#         actor_avg = extract_best_movies(score)
#         # storing actor's average score along with no. of movies (MAX : 5)
#         actor_performance.setdefault(actor_avg, len(score))
#
#     # Score preference based on no. of movies and score
#     count_newbies = 0
#     best_actor_scores = []
#
#     # For loop to accept experienced actors first
#     for i, j in actor_performance.items():
#         # i -> avg score of an actor
#         # j -> no. of movies (MAX 5)
#         if j > 4:
#             best_actor_scores.append(i)
#
#     # Checking if cast list have experienced actors
#     if len(best_actor_scores) > 1:
#         # Scores arranged in descending order
#         best_actor_scores.sort(reverse=True)
#
#         if len(best_actor_scores) >= 3:
#             best_actor_scores = best_actor_scores[:3]
#
#     # if there is lack of experienced actors
#     # newbies are considered
#     if len(best_actor_scores) < 3:
#         for i, j in actor_performance.items():
#             # actors with less experience
#             if j < 4:
#                 best_actor_scores.append(i)
#             if len(best_actor_scores) == 3:
#                 break
#
#     # best 3 actor chosen out of 5
#     p = gd = ok = flop = 0
#     for i in best_actor_scores:
#         if 7.8 <= i:
#             # p counts no. of popular actors
#             p = p + 1
#         elif 7 <= i <= 7.7:
#             # gd counts no. of good actors
#             gd = gd + 1
#         elif 6 <= i < 7:
#             # ok counts no. of average actors
#             ok = ok + 1
#         else:
#             # flop counts no. of flop actors
#             flop = flop + 1
#
#     ws['D' + str(r)].value = p
#     ws['E' + str(r)].value = gd
#     ws['F' + str(r)].value = ok
#     ws['G' + str(r)].value = flop
#
#     r = r + 1
#     movie_name = 'A' + str(r)
#     wb.save('Trial.xlsx')
#
#     count += 1
#     # r = r + 1
#
#
# wb.save('Trial.xlsx')


"""
    Get director average score based on historical data
"""

r = 37
movie_name = 'A' + str(r)
while ws[movie_name].value is not None:
    url = "http://www.omdbapi.com/?apikey=c4779b30&t=" + ws[movie_name].value + "&plot=short"
    response = requests.get(url)
    movie_data = response.json()
    dp = dok = dflop = 0
    if movie_data['Response'] == 'True':
        # A function to extract a list of scores
        # of last 5 movies of a director by passing director name
        scoreOfMovies = director_main_code(movie_data["Director"])

        # scoreOfMovies is score of movies of a director
        # stores the avg of best movies out of 5 or less
        director_avg = extract_best_movies(scoreOfMovies)
        i = director_avg
        if 7 <= i:
            dp = dp + 1
        elif 6 <= i < 7:
            dok = dok + 1
        else:
            dflop = dflop + 1

    ws['H' + str(r)].value = dp
    ws['I' + str(r)].value = dok
    ws['J' + str(r)].value = dflop
    r = r + 1
    movie_name = 'A' + str(r)
    wb.save('Trial.xlsx')

wb.save('Trial.xlsx')



"""
    Code to extract imdb rating of a movie and store in excel sheet 
"""

# url = "http://www.omdbapi.com/?apikey=c4779b30&t=" + ws[row].value + "&plot=short"
# response = r.get(url)
# movie_data = response.json()
# if movie_data['Response'] == 'True':
#     if 'N' not in movie_data['imdbRating']:
#         ws['C' + str(r)].value = float(movie_data['imdbRating'])

"""
    Code to extract average score of actors of a movies searched
    based on historical data of that actor/actress
    
"""

# rel_year_of_movie = 'B' + str(r)
# cast_list = findCast(ws[row].value, ws[rel_year].value)
# score = []
# actor_score = {}
# # if cast List not empty
# score_pref = {}
# actor_performance = {}
# print('Cast : ', cast_list)
# for actor in cast_list:
#     # returns score list of past movies
#     score = cast_score_main_code(actor, ws['B' + str(r)].value)
#     actor_avg = extract_best_movies(score)
#     # storing actor's average score along with no. of movies (MAX : 5)
#     actor_performance.setdefault(actor_avg, len(score))
#
# # Score preference based on no. of movies and score
# count_newbies = 0
# best_actor_scores = []
#
# # For loop to accept experienced actors first
# for i, j in actor_performance.items():
#     # i -> avg score of an actor
#     # j -> no. of movies (MAX 5)
#     if j > 4:
#         best_actor_scores.append(i)
#
# # Checking if cast list have experienced actors
# if len(best_actor_scores) > 1:
#     # Scores arranged in descending order
#     best_actor_scores.sort(reverse=True)
#
#     if len(best_actor_scores) >= 3:
#         best_actor_scores = best_actor_scores[:3]
#
# # if there is lack of experienced actors
# # newbies are considered
# if len(best_actor_scores) < 3:
#     for i, j in actor_performance.items():
#         # actors with less experience
#         if j < 4:
#             best_actor_scores.append(i)
#         if len(best_actor_scores) == 3:
#             break
#
# # best 3 actor chosen out of 5
# p = ok = flop = 0
# for i in best_actor_scores:
#     if 7 <= i:
#         # p counts no. of good actors
#         p = p + 1
#     elif 6 <= i < 7:
#         # ok counts no. of average actors
#         ok = ok + 1
#     else:
#         # flop counts no. of flop actors
#         flop = flop + 1
#
# ws['D' + str(r)].value = p
# ws['E' + str(r)].value = gd
# ws['F' + str(r)].value = ok
# ws['G' + str(r)].value = flop
# r = r + 1
# movie_name = 'A' + str(r)
# wb.save('Trial.xlsx')

