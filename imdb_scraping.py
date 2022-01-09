from openpyxl import workbook, load_workbook
from bs4 import BeautifulSoup
import requests as r
from datetime import date

wb = load_workbook('Trial.xlsx')
ws = wb.active

# wb = load_workbook("DB.xlsx")
# ws = wb.active

row = 128
# url = 'https://www.imdb.com/search/title/?year=2021&title_type=feature&'
start_of_time = ['2021-01-01', '2020-01-01', '2019-01-01', '2018-01-01', '2017-01-01']
end_of_time = ['2021-12-31', '2020-12-31', '2019-12-31', '2018-12-31', '2017-12-31']
start = '1'
# for i, j in zip(start_of_time, end_of_time):
#     print(start_of_time, end_of_time, '\n')
movie_count = 0
k = 0
beginning_movie = ''

while movie_count <= 20:
    i = start_of_time[4]
    j = end_of_time[4]

    # if movie_count % 20 == 0:
    #     k = k + 1
    #     i = start_of_time[k]
    #     j = end_of_time[k]
    # start = str(int(start) + 50)

    url = 'https://www.imdb.com/search/title/?title_type=feature&year=' + i + ',' + j + '&start=' + start + '&ref_=adv_nxt'
    source = r.get(url)
    if source.status_code == 200:
        print('working\n')
        soup = BeautifulSoup(source.text, 'html.parser')
        movies = soup.find('div', class_="lister-list").find_all('div', class_="lister-item")
        print(len(movies))
        for movie in movies:
            Cmovie_name = 'A' + str(row)
            Crelease_year = 'B' + str(row)
            Cimdb = 'C' + str(row)
            print(movie_count, '.', end=" ")
            content = movie.find('div', class_="lister-item-content")
            imdb_score = content.find('div', class_="ratings-bar").find('div', class_="ratings-imdb-rating").find(
                'strong').text
            if 7 <= float(imdb_score) <= 7.7:
                year = content.find('span', class_="lister-item-year").text
                title = content.find('a').text
                print(title, year, imdb_score)
                if movie_count >= 1 and title == beginning_movie:
                    print('\nCondition Reached\n')
                    start = str(int(start) + 50)
                    break

                elif movie_count == 0:
                    beginning_movie = title

                movie_count += 1
                ws[Cmovie_name].value = title
                ws[Crelease_year].value = year[1:-1]
                ws[Cimdb].value = imdb_score
                wb.save("Trial.xlsx")

                row = row + 1

wb.save("Trial.xlsx")
